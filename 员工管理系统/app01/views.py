from django.shortcuts import render, redirect
from app01.utils.pagination import Pagination
from . import models
from .forms import *


def index(request):
    if request.method == 'GET':
        return render(request, 'index.html')


# Create your views here.
"""部门管理"""


def depart_list(request):
    """部门列表"""
    # 1.数据库中获取部门列表
    # queryset=[对象，对象，对象]
    queryset = models.Department.objects.all()

    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    """添加部门"""
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    # 获取用户POST提交的数据
    title = request.POST.get('title')
    # 保存到数据库
    models.Department.objects.create(title=title)

    return redirect('/depart/list/')


def depart_delete(request):
    """删除功能"""
    # 获取id  http://127.00.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')


# 重置
"""    # 先查找传过来的nid
    # nid=request.GET.get('nid')
    # # 表单提交后的数据
    # edit_title=request.POST.get('title')
    # # 查找nid对应的id，并修改里面的title
    # models.Department.objects.filter(id=nid).update(title=edit_title)"""


def depart_edit(request, nid):
    """编辑部门"""
    # GET请求
    if request.method == 'GET':
        # 查找处第一项所有的数据
        row_object = models.Department.objects.filter(id=nid).first()
        ord_title = row_object.title
        # 将旧标题渲染到html的输入框value中
        return render(request, 'depart_edit.html', {"ord_title": ord_title})
    # POST请求
    title = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')


from openpyxl import load_workbook


# 部门列表文件上传
def depart_multi(request):
    """批量上传（excel文件）"""
    from django.core.files.uploadedfile import InMemoryUploadedFile

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.将文件对象传递给workbook，由workbook读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.通过sheet获取到列与行
    # cell = sheet.cell(1, 1)  # (行,列)
    # print(cell.value)  # 值

    # 4.循环获取excel文件中的数据
    for row in sheet.iter_rows(min_row=2):  #iter_rows(min_row=2)       最小行从2开始
        text=row[0].value
        # 将文件中的内容添加到数据库
        exits=models.Department.objects.filter(title=text).exists()
        if not exits:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')


"""用户管理"""


def user_list(request):
    """用户列表"""
    if request.method == 'GET':
        data_list = {}
        search_data = request.GET.get('q', "")  # 有q拿q，没q拿空字符串
        # 判断如果url中的值存在则为True，运行表
        if search_data:
            data_list["mobile__contains"] = search_data

        # 获取数据库中的表
        queryset = models.UserInfo.objects.all()

        """     python语法
           for i in queryset:
            # 日期格式转换
            create_time=i.create_time.strftime('%Y-%m-%d')
            # print(i.get_gender_display())     获取gender字段中的（1，‘男’）中的男数据
            # print(i.depart.title) 根据id自动去关联的表中获取哪一行数据depart对象。"""

        page_objects = Pagination(request, queryset)
        # 传递的所有数据
        content = {'queryset': page_objects.queryset,  # 分页完成的数据
                   "search_data": search_data,
                   "page_string": page_objects.html(),  # # 获取定义方法的返回值   页码
                   "all_count": page_objects.total_page_count
                   }
        return render(request, 'user_list.html', content)
    # POST请求
    go_page = request.POST.get('go_page', 1)

    if go_page:
        if 1 <= int(go_page) <= 2:  # 判断是否小于或大于总数居
            return redirect(f'/user/list/?page={go_page}')
    return redirect(f'/user/list/')


def user_add(request):
    """普通增加用户"""
    if request.method == 'GET':
        queryset = models.Department.objects.all()
        content = {
            'gender_choices': models.UserInfo.gender_choices,
            'depart_list': models.Department.objects.all()

        }
        return render(request, 'user_add.html',
                      {'departs': content.get('depart_list'), "content": content.get('gender_choices')})
    name = request.POST.get('name')
    password = request.POST.get('password')
    age = request.POST.get('age')
    account = request.POST.get('account')
    create_time = request.POST.get('create_time')
    gender = request.POST.get('gender')
    depart_id = request.POST.get('depart')
    print(create_time)
    models.UserInfo.objects.create(name=name, password=password, age=age, account=account, create_time=create_time,
                                   gender=gender, depart_id=depart_id)
    return redirect('/user/list/')


# ---------------------------modelform---------------------------


def user_add_modelform(request):
    """modelform验证用户添加"""

    if request.method == 'GET':
        form = UserModelForm()
        return render(request, 'user_add_modelform.html', {'form': form})

    # POST提交的数据，进行校验
    form = UserModelForm(data=request.POST)
    if form.is_valid():
        # 数据成功
        # {'name': '1', 'password': '1', 'age': 1, 'account': Decimal('0'), 'create_time': datetime.datetime(2023, 5, 12, 0, 0, tzinfo=<UTC>), 'gender': 0, 'depart': <Department: 销售部>}
        # print(form.cleaned_data)
        form.save()  # 提交到UserModelForm类下定义的model变量中
        return redirect('/user/list/')

    # 失败(显示错误信息)
    return render(request, 'user_add_modelform.html', {'form': form})


def user_delete(request, nid):
    """删除用户"""
    models.UserInfo.objects.filter(id=nid).delete()

    return redirect('/user/list/')


def user_edit(request, nid):
    """编辑用户"""

    row_object = models.UserInfo.objects.filter(id=nid).first()
    if request.method == 'GET':
        # 根据id去数据库中获取需要编辑的那一行数据

        form = UserModelForm(instance=row_object)
        return render(request, 'user_edit.html', {'form': form})

    # 1.获取用户提交的数据
    form = UserModelForm(request.POST, instance=row_object)  # 4.stance=row_object
    # 2.校验数据是否为空或合法
    if form.is_valid():
        # 3.提交至数据库，当前为添加，需要指定到是哪一行数据才能是更新数据 instance=row_object

        form.save()
        # 也可以form.instance.字段名=123
        return redirect('/user/list/')
    return render(request, 'user_edit.html', {'form': form})


# ---------------------------modelform---------------------------


def goods_phone_list(request):
    """靓号列表"""

    # 搜索
    # 获取url地址的值
    now_page = 0

    if request.method == 'GET':
        data_list = {}
        search_data = request.GET.get('q', "")  # 有q拿q，没q拿空字符串
        # 判断如果url中的值存在则为True，运行表
        if search_data:
            data_list["mobile__contains"] = search_data

        # 分页
        queryset = models.PrettyNum.objects.filter(**data_list).order_by('-level')

        page_objects = Pagination(request, queryset)  # 将参数传递给组件

        # 传递的所有数据
        content = {'queryset': page_objects.queryset,  # 分页完成的数据
                   "search_data": search_data,
                   "page_string": page_objects.html(),  # # 获取定义方法的返回值   页码
                   "all_count": page_objects.total_page_count
                   }
        return render(request, 'pretty_list.html', content)
    # POST请求
    go_page = request.POST.get('go_page', 1)

    if go_page:
        if 1 <= int(go_page) <= 55:  # 判断是否小于或大于总数居
            return redirect(f'/pretty/list/?page={go_page}')
    return redirect(f'/pretty/list/')


def pretty_add(request):
    """添加靓号"""

    if request.method == 'GET':
        # 实例化
        form = PrettyModelForm()
        # 字段渲染到页面
        return render(request, 'pretty_add.html', {'form': form})
    # 请求回来的数据在PrettyModelForm中进行校验
    form = PrettyModelForm(data=request.POST)

    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')
    return render(request, 'pretty_add.html', {'form': form})


def pretty_edit(request, nid):
    """编辑靓号"""
    row_object = models.PrettyNum.objects.filter(id=nid).first()
    if request.method == 'GET':
        # instance实例化对象为row_object
        form = PrettyEditModelForm(instance=row_object)
        return render(request, 'pretty_edit.html', {'form': form})
    # 有instance=row_object才是在原来的基础上更新数据，否则是添加新数据
    form = PrettyEditModelForm(request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')

    return render(request, 'pretty_edit.html', {'form': form})


def pretty_delete(request, nid):
    """删除靓号"""
    models.PrettyNum.objects.filter(id=nid).delete()
    page = request.GET.get('page')

    return redirect(f'/pretty/list/?page={page}')


# ----------------------------管理员----------------------------

def admin_list(request):
    """管理员列表"""

    # 检查用户是否登录，没有登录则跳转到登录页码，登陆成功后显示信息
    # 用户发来请求，获取cookie随机字符串，拿着随机字符串看看session有没有

    # 获取session中的字典
    # username = request.session.get('info')
    # # 如果没有登录则跳转到登录页面
    # if not username:
    #     return redirect('/login/')

    if request.method == 'GET':
        search_data = request.GET.get('q', '')
        # 存放搜索数据的字典
        search_list = {}
        if search_data:
            # 包含search_data的用户名
            search_list['username__contains'] = search_data
        # 查找
        queryset = models.Admin.objects.filter(**search_list)
        page_object = Pagination(request, queryset)
        content = {
            "queryset": page_object.queryset,
            "page_string": page_object.html(),
            "all_count": page_object.total_page_count,
        }
        return render(request, 'admin_list.html', content)


def admin_add(request):
    """添加管理员"""
    title = '管理员'
    if request.method == 'GET':
        form = AdminModelForm()

        content = {
            "name": title,
            "form": form
        }
        return render(request, 'change_add.html', content)
    # POST
    form = AdminModelForm(data=request.POST)
    if form.is_valid():  # 判断有效则
        # {'username': '111', 'password': '111', 'confirm_password': 14124}      #confirm_password为空，是因为钩子函数没有返回值
        form.save()

        return redirect("/admin/list/")

    content = {
        "name": title,
        "form": form
    }
    return render(request, 'change_add.html', content)


def admin_edit(request, nid):
    """编辑管理员"""
    title = '管理员'
    # 对象 /None
    now_object = models.Admin.objects.filter(id=nid).first()  # 判断传递的id是否在数据库存在
    if request.method == 'GET':

        if not now_object:  # 不存在则
            return redirect("/admin/list/")
        # 只显示用户名
        form = AdminEditModelForm(instance=now_object)

        content = {
            "name": title,
            "form": form
        }
        return render(request, 'change_add.html', content)
    form = AdminEditModelForm(data=request.POST, instance=now_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    content = {
        "name": title,
        "form": form
    }
    return render(request, 'change_add.html', content)


def admin_delete(request, nid):
    """删除管理员"""
    models.Admin.objects.filter(id=nid).delete()
    return redirect("/admin/list/")


def admin_reset(request, nid):
    """重置密码"""
    now_object = models.Admin.objects.filter(id=nid).first()
    title = f'重置密码 -{now_object.username}'
    if not now_object:
        return redirect("/admin/list/")
    if request.GET == 'GET':
        form = AdminResetModelForm
        content = {
            "name": title,
            "form": form
        }

        return render(request, 'reset.html', content)

    # POST
    form = AdminResetModelForm(data=request.POST, instance=now_object)
    if form.is_valid():
        # print(form.cleaned_data)
        form.save()
        return redirect("/admin/list/")

    content = {
        "name": title,
        "form": form
    }

    return render(request, 'reset.html', content)
